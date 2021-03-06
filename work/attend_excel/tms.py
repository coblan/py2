# -*- encoding:utf8 -*-
"""
模拟TMS数据
直接运行该文件，会运行geninfo()函数，其目的是从record,report提取雇员的信息，如雇员号，考勤号等，将模型EmployModel。最后保存为sqlite3文件
该文件被别的模块调用时，主要调用TmsData.empoyee()生成器，其目的是罗列出雇员信息。
EmployModel 当前雇员模型

"""
import openpyxl,pickle
from orm.model import Model
from orm.fields import CharField
import sqlite3
import os
BASE_DIR = os.path.dirname(__file__)


class EmployModel(Model):
    "雇员的信息，当前信息来源于从excel提取出的固定信息"
    empid = CharField(default="")
    kao_number = CharField(default='')
    name = CharField(default='')
    pname =CharField(default='')
    status = CharField(default='')
    work_type = CharField(default='')
    work_shift = CharField(default='')
    #expect_days = Field()
    startjob = CharField(default='')
    startmokie = CharField(default='')
    
#EmployModel.connection(sqlite3.connect("tmpfiles/employee"))

class TmsData(object):
    
    @staticmethod
    def loadFromExcel():
        EmployModel.connection(sqlite3.connect(":memory:"))
        EmployModel.create()
        wb=openpyxl.load_workbook(os.path.join(BASE_DIR,"employee.xlsx"))
        ws= wb.active
        head=True
        for i in ws.rows:
            if head:  
                head =False
                continue  
            ls = [j.value for j in i]
            EmployModel(empid=ls[0],kao_number=str(ls[1]),name=ls[2],pname=ls[3],status=ls[4],work_type=ls[5],
                        work_shift=ls[6],startjob=ls[7],startmokie=ls[8]).save()
        EmployModel.commit()
        
    @staticmethod
    def employee():
        "罗列出雇员"
        for p in EmployModel.select():
            #if p.empid =='AE1774':
                #p.kao_number = 1157
                #p.save()
            #elif p.empid == 'AE1776':
                #p.kao_number= 1158 
                #p.save()
            yield p  
    
    @staticmethod
    def from_kaonumber(kao_number):
        for p in EmployModel.select("WHERE kao_number='%s'"%kao_number):
            return p
    @staticmethod
    def kao2workshift(kao_number):
        person = TmsData.from_kaonumber(kao_number)
        if person:
            return person.work_shift
        else:
            return ''
        
def get_workshift(empid):
    for i in EmployModel.select("WHERE empid='%s'"%empid):
        return i.work_shift

def get_worktype(empid):
    for i in EmployModel.select("WHERE empid='%s'"%empid):
        return i.work_type

def attend_num_to_empid(kao_num):
    for i in EmployModel.select("WHERE kao_number='%s'"%kao_num):
        return i.empid

def geninfo():
    "直接运行该函数，将解析xlsx与xls文件，将生成一个列表，并保存到硬盘，目的是将empID同kao_number对应起来。"

    from attend_reader import Reader
    
    conn = sqlite3.connect(':memory:')
    Reader.read03(conn, r"D:\work\attendance\attendance record.xls")
    cursor = conn.cursor()
    Raw_Record.connection(conn)
    
    '从excel提取员工信息，保存为pickle格式，作为测试用'

    EmployModel.connection(conn)
    EmployModel.create()
    
    wb = openpyxl.load_workbook(r"D:\work\attendance\MK_HR_Attendance Report_Cecilia_201510.xlsx")
    ws = wb['report']
    out =[]
    head=True
    for i in ws.rows:
        if head:
            head =False
            continue
        ls = [j.value for j in i]
        kao_number='000'
        for i in Raw_Record.select("WHERE name='%s'"%ls[1]):
            kao_number = i.kao_number
            break
        EmployModel(empid=ls[0],kao_number=kao_number,name=ls[1],pname=ls[2],status=ls[3],work_type=ls[4],work_shift=ls[5]).save()

    EmployModel.commit()
    
    wb2 = openpyxl.load_workbook(r"D:\work\attendance\MK_HR_Leave Record2015_TEST.xlsx")
    ws2 = wb2["Annual Leave"]
    for i in ws2.rows:
        ls = [j.value for j in i]
        empid = ls[0]
        for i in EmployModel.select("WHERE empid='%s'"%empid):
            try:
                i.startjob = ls[5].strftime("%Y/%m/%d")
            except:
                i.startjob=""
            try:
                i.startmokie = ls[6].strftime("%Y/%m/%d")
            except:
                i.startmokie =""
            i.save()
            
    EmployModel.commit()
    
    wb2= openpyxl.Workbook()
    ws = wb2.active
    ws.append(["empid","kao qin number","name","pin yin name","status","work type","work shift","work experience","mokie experience"])
    for row in EmployModel.select("ORDER BY empid"):
        assert isinstance(row,EmployModel)
        ws.append([row.empid,row.kao_number,row.name,row.pname,row.status,row.work_type,row.work_shift,row.startjob,row.startmokie])
    wb2.save("employee.xlsx")


if __name__ !="__main__":
    TmsData.loadFromExcel()

if __name__ =='__main__':
    # from attend_reader import Raw_Record
    # geninfo()
    TmsData.loadFromExcel()
    print( get_workshift('AE1776'))