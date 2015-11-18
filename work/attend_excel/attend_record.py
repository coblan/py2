# -*- encoding:utf8 -*-
from mytime import Time
from openpyxl.styles import PatternFill,Style,Color
from settings import *

class Record(object):
    """
    0:kao_number,8:team_late,9:workspan
    10:person_late
    11:over_time
    """
    def __init__(self,conn):
        self.conn =conn
        
    def parse(self):
        record_cursor= self.conn.cursor()
        record_cursor.execute('''CREATE TABLE record
                             (kao_number text, name text, department text, date text, workstart text, workleave text,
                             note text, sub_sequence text,team_late text,workspan text,person_late text,over_time text)''') 
        raw_cursor = self.conn.cursor()
        for row in raw_cursor.execute("""SELECT * FROM raw_record"""):
            personId=row[0]

            # 输出字段
            kao_number= row[0]
            name = row[1]
            department = row[2]
            date = row[3]
            workstart = Time.strptime(row[4])
            workleave = Time.strptime(row[5])
            note = ''
            sub_sequence = self.get_sub_sequence(workstart,personId)
            team_late = self.get_team_late(workstart,personId)
            workspan= (workleave- workstart-Time(1) ) if isinstance(workleave,Time) else ""
            person_late = self.get_person_late(workstart,personId)
            over_time= self.get_over_time(workstart,workleave,personId)

            outrow = (kao_number,name,department,date,str(workstart),str(workleave),note,sub_sequence,str(team_late),str(workspan),str(person_late),str(over_time))
            record_cursor.execute("INSERT INTO record VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"%outrow)

        self.conn.commit()   
        return self
    
    def gen07(self,ws):
        ws.title='record'
        ws.append([u"考勤号码",u"姓名",u"部门",u"日期",u"上班时间",u"下班时间",u'Note',u'sub-sequence',u'迟到时长-团队',u'工作时长',u'迟到时长-个人出勤率',u'加班时长'])
        
        cnt=1
        record_cursor = self.conn.cursor()
        #record_cursor.execute('''CREATE TABLE record
                     #(kao_number text, name text, department text, date text, workstart text, workleave text,
                     #note text, sub_sequence text,team_late text,workspan text,person_late text,over_time text)''')   
        #raw_cursor = self.data.cursor()
        for row in record_cursor.execute("""SELECT * FROM record"""):
            cnt+=1
            personId=row[0]

            # 输出字段
            workstart = Time.strptime(row[4])
            workleave = Time.strptime(row[5])


            #判断添加cell颜色，要等待写入2007后，才能添加颜色，见【1】
            shang=self.shang_ban_color(workstart)
            xia =self.xia_ban_color(workleave)

            # 整理格式准备写入excel2007
            #outrow = [kao_number,name,department,date,str(workstart),str(workleave),note,sub_sequence,str(team_late),str(workspan),str(person_late),str(over_time)]
            #record_cursor.execute("INSERT INTO record VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"%tuple(outrow))
            ws.append(row)  

            #添加颜色【1】
            if shang:
                ws['E'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=shang)
            if xia:
                ws['F'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=xia)      
        
    def get_team_late(self,start,personId):
        'callback:返回迟到团队时间，Time对象'
        if isinstance(start,Time):
            spot = Time(8,45)
            late = start - spot
            return late
        else:
            return ''

    def get_sub_sequence(self,workstart,personId):
        if workstart and isinstance(workstart,Time):
            for k,v in late.items():
                if v[0]<=workstart<=v[1]:
                    return k
        return ''

    def get_person_late(self, workstart,personId):
        late = self.get_sub_sequence(workstart,personId)
        if  late== 'late2':
            return workstart-Time(8,45)
        elif late=='late3':
            late = workstart -Time(8,45)
            return late*2
        else:
            return Time(0)

    def get_over_time(self,workstart, workleave,personId):
        if isinstance(workstart,Time):
            return workleave-Time(20)
        else:
            return Time(0)


    def shang_ban_color(self,workstart):
        if not workstart:
            return
        late = self.get_sub_sequence(workstart, personId=None)
        if late=='late3':
            return 'FFFFFF00'
        elif late == 'late2':
            return 'FFA4D3EE'
        elif late == 'late1':
            return 'FFFFB5C5'

    def xia_ban_color(self,workleave):
        if workleave:
            return        