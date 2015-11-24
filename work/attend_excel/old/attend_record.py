# -*- encoding:utf8 -*-
# 生成考勤记录统计
#
#
from mytime import Time
from openpyxl.styles import PatternFill,Style,Color
from settings import *
from attend_reader import Raw_Record
from orm import Model,Field
from tms import TmsData
from datetime import datetime,timedelta,date


class RecordModel(Model): 
    "整理后的考勤记录数据表"
    kao_number=Field()
    name = Field()
    department = Field()
    date = Field()
    workstart = Field()
    workleave = Field()
    note = Field()
    sub_sequence = Field()
    late_team = Field()
    workspan = Field()
    late_person=Field()
    overtime =Field()


class Record(object):
    def __init__(self):
        self.crtperson = None
        
    
    def gen_table(self,conn):
        "在conn中生成处理的考勤记录表,表模型见RecordModel"
        RecordModel.connection(conn)
        RecordModel.create()
        
        for row in Raw_Record.select():
            self.crtperson = row
            
            workstart = Time.strptime(row.workstart)
            workleave = Time.strptime(row.workleave)           
            # 输出字段
            RecordModel(
                kao_number= row.kao_number,
                name = row.name,
                department = row.department,
                date = datetime.strptime(row.date,"%Y/%m/%d").date().strftime("%Y/%m/%d"),
                workstart = workstart,
                workleave = workleave,
                note = '',
                sub_sequence = self.get_sub_sequence(),
                late_team = self.get_late_team(),
                workspan= (workleave- workstart-Time(1) ) if isinstance(workleave,Time) else "",
                late_person = self.get_late_person(),
                overtime= self.get_over_time(),
                ).save()
 
        RecordModel.commit()
        
        overtimelist = []
        for p in RecordModel.select():
            assert isinstance(p,RecordModel)
            overtime = Time.strptime(p.overtime)
            if overtime!=Time(0):
                overtimelist.append(p)
        
        for p in overtimelist:
            date = datetime.strptime(p.date,"%Y/%m/%d").date()
            self.crtperson = p
            if date.weekday()!=4:
                lastday = date + timedelta(days=1)
                laststr = lastday.strftime("%Y/%m/%d")
                for i in RecordModel.select("WHERE date='%s' AND kao_number='%s'"%(laststr,p.kao_number)):
                    i.late_team= self.get_late_team(tic=Time(10))
                    i.late_person = self.get_late_person(tic=Time(10))
                    i.sub_sequence = self.get_sub_sequence(Time(10))
                    i.save()
                    break
                
    
    def gen07(self,ws):
        ws.title='record'
        ws.append([u"考勤号码",u"姓名",u"部门",u"日期",u"上班时间",u"下班时间",u'Note',u'sub-sequence',u'迟到时长-团队',u'工作时长',u'迟到时长-个人出勤率',u'加班时长'])
        
        cnt=1
        for row in RecordModel.select():
            cnt+=1
            assert isinstance(row,RecordModel)
            self.crtperson=row

            #判断添加cell颜色，要等待写入2007后，才能添加颜色，见【1】
            shang=self.shang_ban_color()
            xia =self.xia_ban_color()

            # 整理格式准备写入excel2007
            outrow=[row.kao_number,row.name,row.department,row.date,row.workstart,row.workleave,row.note,row.sub_sequence,row.late_team,row.workspan,row.late_person,row.overtime]
            ws.append(outrow)  

            #添加颜色【1】
            if shang:
                ws['E'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=shang)
            if xia:
                ws['F'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=xia)      
        
        
    def get_late_team(self,tic=Time(8,30)):
        'callback:返回迟到团队时间，Time对象'
        workstart = Time.strptime(self.crtperson.workstart)
        if isinstance(workstart,Time):
            late = workstart - (tic+Time(0,15))
            return late
        
        return ''

    def get_sub_sequence(self,tic=Time(8,30)):
        workstart = Time.strptime(self.crtperson.workstart)
        if isinstance(workstart,Time):
            if tic< workstart <=tic+Time(0,15):
                return 'late1'
            elif tic+Time(0,15) <= workstart <tic+Time(1):
                return 'late2'
            elif tic+Time(1) <= workstart:
                return 'late3'
            #for k,v in lateLevel.items():
                #if v[0]<=workstart<=v[1]:
                    #return k
        
        return ''

    def get_late_person(self,tic=Time(8,30)):
        workstart = Time.strptime(self.crtperson.workstart)
        if isinstance(workstart,Time):
            if tic+Time(0,15) <=workstart <tic+Time(1):
                return workstart-(tic+Time(0,15) )
            elif tic+Time(1)<= workstart:
                return (workstart- (tic+Time(1)))*2
       
        return Time(0)
        

    def get_over_time(self):
        #workstart = Time.strptime(self.crtperson.workstart)
        workleave = Time.strptime(self.crtperson.workleave)
        if isinstance(workleave,Time):
            return workleave-Time(20)
        else:
            return Time(0)


    def shang_ban_color(self):
        late = self.crtperson.sub_sequence
        if late=='late3':
            return 'FFFFFF00'
        elif late == 'late2':
            return 'FFA4D3EE'
        elif late == 'late1':
            return 'FFFFB5C5'

    def xia_ban_color(self):
        pass       