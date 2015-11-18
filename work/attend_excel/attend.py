# -*- encoding:utf8 -*-

import sqlite3
from datetime import time
from openpyxl import Workbook
#from openpyxl.styles import PatternFill,Style,Color
from settings import *
#from mytime import Time
from attend_reader import Reader
from attend_record import Record
from attend_report import Report

from tms import TmsData

class Attendence(object):
    def __init__(self):
        self.conn = sqlite3.connect(':memory:')
        
    def read03(self,path):
        Reader.read03(self.conn,path)
        #att = Attendence()
        #rawdata = xlrd.open_workbook(path,encoding_override='gbk')
        
        #table = rawdata.sheets()[0]
        #nrows = table.nrows
        #c = att.data.cursor()
        #c.execute('''CREATE TABLE raw_record
             #(kao_number text, name text, department text, date text, workstart text, workleave text)''')

        #for i in range(1, nrows):
            #start,end=find_min_max(table.row_values(i)[4])
            #c.execute("""INSERT INTO raw_record VALUES ('%s','%s','%s','%s','%s','%s')"""%(table.row_values(i)[0],table.row_values(i)[1],table.row_values(i)[2],table.row_values(i)[3],start,end))
        #att.data.commit()
        #return att
    def process(self):
        self.record = Record(self.conn)
        self.record.parse()
        self.report = Report(self.conn,TmsData())
        self.report.report()
    def gen07(self):
        self.wb = Workbook()
        ws= self.wb.active 
        self.record.gen07(ws)
        #ws.title='record'
        #ws.append([u"考勤号码",u"姓名",u"部门",u"日期",u"上班时间",u"下班时间",u'Note',u'sub-sequence',u'迟到时长-团队',u'工作时长',u'迟到时长-个人出勤率',u'加班时长'])
        #cnt=1
        #record_cursor = self.data.cursor()
        #record_cursor.execute('''CREATE TABLE record
                     #(kao_number text, name text, department text, date text, workstart text, workleave text,
                     #note text, sub_sequence text,team_late text,workspan text,person_late text,over_time text)''')   
        #raw_cursor = self.data.cursor()
        #for row in raw_cursor.execute("""SELECT * FROM raw_record"""):
            #cnt+=1
            #personId=row[0]
            
            ## 输出字段
            #kao_number= row[0]
            #name = row[1]
            #department = row[2]
            #date = row[3]
            #workstart = Time.strptime(row[4])
            #workleave = Time.strptime(row[5])
            #note = ''
            #sub_sequence = self.get_sub_sequence(workstart,personId)
            #team_late = self.get_team_late(workstart,personId)
            #workspan= (workleave- workstart-Time(1) ) if isinstance(workleave,Time) else ""
            #person_late = self.get_person_late(workstart,personId)
            #over_time= self.get_over_time(workstart,workleave,personId)
            
            ##判断添加cell颜色，要等待写入2007后，才能添加颜色，见【1】
            #shang=self.shang_ban_color(workstart)
            #xia =self.xia_ban_color(workleave)
            
            ## 整理格式准备写入excel2007
            #outrow = [kao_number,name,department,date,str(workstart),str(workleave),note,sub_sequence,str(team_late),str(workspan),str(person_late),str(over_time)]
            #record_cursor.execute("INSERT INTO record VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"%tuple(outrow))
            #ws.append(outrow)  
            
            ##添加颜色【1】
            #if shang:
                #ws['E'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=shang)
            #if xia:
                #ws['F'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=xia)
        #self.data.commit()
        ws2=self.wb.create_sheet()
        self.report.gen07(ws2)
        
    def save(self,path):
        self.wb.save(path)
        

if __name__ == '__main__':
    testpath = r"D:\work\attendance\attendance record.xls"
    testsavepath="tmpfiles/test.xlsx"
    
    attend = Attendence()
    attend.read03(testpath)
    attend.process()
    attend.gen07()
    attend.save(testsavepath)
    #Attendence\
        #.read03(testpath)\
        #.gen07()\
        #.save(testsavepath)
