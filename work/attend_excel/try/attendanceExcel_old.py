# -*- encoding:utf8 -*-

import xlrd,sqlite3,att_report
from datetime import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Style,Color

late={
    'late1':["8:30","8:45"],
    'late2':["8:46","9:30"],
    'late3':["9:31","17:30"]
}

class Time(object):
    '时间运算，转换。'
    def __init__(self,hour,minute=0):
        self.hour=hour
        self.minute =minute
    
    @staticmethod
    def strptime(str_):
        if not str_:
            return ''
        tm = Time(0)
        ls= str_.split(":")
        if ls:
            tm.hour = int(ls[0])
            tm.minute = int(ls[1]) 
        return tm
    
    def __sub__(self,other):
        tm1 = self.hour*60 + self.minute
        tm2 = other.hour*60 +other.minute
        span = tm1 -tm2
        if span <0:
            span = 0
        hour = span/60
        minute = span%60
        return Time(hour,minute)
    
    def __cmp__(self,other):
        if self.hour == other.hour:
            if self.minute<other.minute:
                return -1
            elif self.minute == other.minute:
                return 0
            elif self.minute>other.minute:
                return 1
        elif self.hour>other.hour:
            return 1
        elif self.hour < other.hour:
            return -1
    
    def __mul__(self,number):
        hour = self.hour*number
        minute = self.minute*number
        total = hour*60 +minute
        hour = total/60
        minute = total%60
        return Time(hour,minute)
    
    def __str__(self):
        return '%s:%02d:00'%(self.hour,self.minute)
    
    def __add__(self,other):
        hour = self.hour+other.hour
        minute = self.minute+other.minute
        hour = hour + minute/60
        minute = minute%60
        return Time(hour,minute)
    
    def to_stdtime(self):
        hour =self.hour%24
        return time(hour,self.minute)
        


for k ,v in late.items():
    late[k]=[Time.strptime(str_) for str_ in v]

def find_min_max( timeList):
    timeList = timeList.strip()
    if timeList:
        ls = timeList.split(u' ')
        if ls:
            ls = [Time.strptime(str_) for str_ in ls]
            ls.sort()
            return ls[0],ls[-1]
    
    return '',''
        

class Attendence(object):
    def __init__(self):
        self.data = sqlite3.connect(':memory:')
        
    @staticmethod
    def read03(path):
        att = Attendence()
        rawdata = xlrd.open_workbook(path,encoding_override='gbk')
        
        table = rawdata.sheets()[0]
        nrows = table.nrows
        c = att.data.cursor()
        c.execute('''CREATE TABLE raw_record
             (kao_number text, name text, department text, date text, workstart text, workleave text)''')

        for i in range(1, nrows):
            start,end=find_min_max(table.row_values(i)[4])
            c.execute("""INSERT INTO raw_record VALUES ('%s','%s','%s','%s','%s','%s')"""%(table.row_values(i)[0],table.row_values(i)[1],table.row_values(i)[2],table.row_values(i)[3],start,end))
        att.data.commit()
        return att
        
    def gen07(self):
        self.wb = Workbook()
        ws= self.wb.active 
        ws.title='record'
        ws.append([u"考勤号码",u"姓名",u"部门",u"日期",u"上班时间",u"下班时间",u'Note',u'sub-sequence',u'迟到时长-团队',u'工作时长',u'迟到时长-个人出勤率',u'加班时长'])
        cnt=1
        record_cursor = self.data.cursor()
        record_cursor.execute('''CREATE TABLE record
                     (kao_number text, name text, department text, date text, workstart text, workleave text,
                     note text, sub_sequence text,team_late text,workspan text,person_late text,over_time text)''')   
        raw_cursor = self.data.cursor()
        for row in raw_cursor.execute("""SELECT * FROM raw_record"""):
            cnt+=1
            personId=row[0]
            
            # 输出字段
            kao_number= row[0]
            name = row[1]
            department = row[2]
            date = row[3]
            workstart = Time.strptime(row[4])
            workleave = Time.strptime(row[5])
            note = ''
            sub_sequence = self.get_sub_sequence(workstart,personId)
            team_late = self.get_team_late(workstart,personId)
            workspan= (workleave- workstart-Time(1) ) if isinstance(workleave,Time) else ""
            person_late = self.get_person_late(workstart,personId)
            over_time= self.get_over_time(workstart,workleave,personId)
            
            #判断添加cell颜色，要等待写入2007后，才能添加颜色，见【1】
            shang=self.shang_ban_color(workstart)
            xia =self.xia_ban_color(workleave)
            
            # 整理格式准备写入excel2007
            outrow = [kao_number,name,department,date,str(workstart),str(workleave),note,sub_sequence,str(team_late),str(workspan),str(person_late),str(over_time)]
            record_cursor.execute("INSERT INTO record VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"%tuple(outrow))
            ws.append(outrow)  
            
            #添加颜色【1】
            if shang:
                ws['E'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=shang)
            if xia:
                ws['F'+str(cnt)].fill=PatternFill(fill_type='solid', start_color=xia)
        self.data.commit()
        
        att_report.Report(self.data,self.wb).gen07()
        return self
    
    def get_team_late(self,start,personId):
        'callback:返回迟到团队时间，Time对象'
        if isinstance(start,Time):
            spot = Time(8,45)
            late = start - spot
            return late
        else:
            return ''
    
    def get_sub_sequence(self,workstart,personId):
        if workstart and isinstance(workstart,Time):
            for k,v in late.items():
                if v[0]<=workstart<=v[1]:
                    return k
        return ''
    
    def get_person_late(self, workstart,personId):
        late = self.get_sub_sequence(workstart,personId)
        if  late== 'late2':
            return workstart-Time(8,45)
        elif late=='late3':
            late = workstart -Time(8,45)
            return late*2
        else:
            return Time(0)
    
    def get_over_time(self,workstart, workleave,personId):
        if isinstance(workstart,Time):
            return workleave-Time(20)
        else:
            return Time(0)

    
    def shang_ban_color(self,workstart):
        if not workstart:
            return
        late = self.get_sub_sequence(workstart, personId=None)
        if late=='late3':
            return 'FFFFFF00'
        elif late == 'late2':
            return 'FFA4D3EE'
        elif late == 'late1':
            return 'FFFFB5C5'

    def xia_ban_color(self,workleave):
        if workleave:
            return
        
    def save(self,path):
        self.wb.save(path)
        
#data = xlrd.open_workbook(r"D:\work\attendance\attendance record.xls",encoding_override='gbk')
#table = data.sheets()[0] 
#for i in table.row_values(1):
    #print(repr(i))


#wb = Workbook()

## grab the active worksheet
#ws = wb.active

## Data can be assigned directly to cells
#ws['A1'] = 42

## Rows can also be appended
#ws.append([1, 2, 3])

## Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

##ws['B2'].style = Style(fill=PatternFill(patternType='solid', fgColor=Color('FF000000'))) 
#ws['B2'].style = Style(fill=PatternFill(fill_type='solid', start_color='FFFF0000',end_color='FFFF0000')) 

## Save the file
#wb.save("sample.xlsx")

if __name__ == '__main__':
    testpath = r"D:\work\attendance\attendance record.xls"
    testsavepath="test.xlsx"
    
    Attendence\
        .read03(testpath)\
        .gen07()\
        .save(testsavepath)
