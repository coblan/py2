# -*- encoding:utf8 -*-

import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Style,Color
data = xlrd.open_workbook(r"D:\work\attendance\attendance record.xls",encoding_override='gbk')
table = data.sheets()[0] 
for i in table.row_values(0):
    print(i)


wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

#ws['B2'].style = Style(fill=PatternFill(patternType='solid', fgColor=Color('FF000000'))) 
ws['B2'].style = Style(fill=PatternFill(fill_type='solid', start_color='FFFF0000',end_color='FFFF0000')) 

# Save the file
wb.save("sample.xlsx")